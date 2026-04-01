"""
CONVERSOR CSV → EXCEL
Transforma um CSV em planilha Excel formatada, com resumo e gráfico.

USO:
  python conversor.py dados.csv
  python conversor.py dados.csv --saida relatorio

INSTALAR:
  pip install pandas openpyxl
"""

import pandas as pd
import argparse
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


def converter(caminho_csv, nome_saida=None):
    caminho = Path(caminho_csv)
    saida   = caminho.parent / f"{nome_saida or caminho.stem}.xlsx"

    print(f"\n  Lendo: {caminho.name}")
    df = pd.read_csv(caminho)
    print(f"  {len(df)} linhas, {len(df.columns)} colunas")

    # Passo 1: pandas grava o Excel bruto na aba "Dados"
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados", index=False)

    # Passo 2: openpyxl reabre para formatar
    # (pandas não expõe toda a formatação do openpyxl diretamente)
    wb = load_workbook(saida)
    ws = wb["Dados"]

    formatar_dados(ws, df)
    adicionar_resumo(wb, df)
    adicionar_grafico(wb, df)

    wb.save(saida)
    print(f"  ✅ Salvo: {saida}\n")


def formatar_dados(ws, df):
    """Cabeçalho azul, linhas alternadas, colunas com largura automática."""

    fill_cabecalho = PatternFill("solid", fgColor="2F5496")
    fill_alternado = PatternFill("solid", fgColor="DCE6F1")
    fonte_branca   = Font(bold=True, color="FFFFFF")

    # Formata a linha do cabeçalho
    for cel in ws[1]:
        cel.fill      = fill_cabecalho
        cel.font      = fonte_branca
        cel.alignment = Alignment(horizontal="center")

    # Formata as linhas de dados com cor alternada
    for i, linha in enumerate(ws.iter_rows(min_row=2), start=2):
        for cel in linha:
            if i % 2 == 0:
                cel.fill = fill_alternado

    # Ajusta a largura de cada coluna ao conteúdo
    for idx, coluna in enumerate(ws.columns, start=1):
        maior = max(len(str(cel.value or "")) for cel in coluna)
        ws.column_dimensions[get_column_letter(idx)].width = min(maior + 4, 50)

    # Trava o cabeçalho (fica visível ao rolar)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def adicionar_resumo(wb, df):
    """Cria a aba Resumo com estatísticas das colunas numéricas."""
    ws = wb.create_sheet("Resumo")

    ws["A1"] = "Total de linhas"
    ws["B1"] = len(df)
    ws["A2"] = "Total de colunas"
    ws["B2"] = len(df.columns)

    ws["A4"] = "Estatísticas (colunas numéricas)"
    ws["A4"].font = Font(bold=True)

    # df.describe() calcula contagem, média, desvio padrão, mín, máx etc.
    # .T transpõe para que cada coluna numérica vire uma linha da tabela
    desc = df.describe().T.reset_index()

    for col_idx, nome in enumerate(desc.columns, start=1):
        cel = ws.cell(row=5, column=col_idx, value=nome)
        cel.font = Font(bold=True)
        cel.fill = PatternFill("solid", fgColor="2F5496")
        cel.font = Font(bold=True, color="FFFFFF")

    for row_idx, linha in enumerate(desc.itertuples(index=False), start=6):
        for col_idx, valor in enumerate(linha, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)


def adicionar_grafico(wb, df):
    """Gráfico de barras da primeira coluna numérica na aba Gráfico."""
    numericas = df.select_dtypes(include="number").columns
    if len(numericas) == 0:
        return

    ws_dados   = wb["Dados"]
    ws_grafico = wb.create_sheet("Gráfico")

    # Índice da coluna no Excel (começa em 1)
    idx_col   = list(df.columns).index(numericas[0]) + 1
    max_linha = min(21, len(df) + 1)

    grafico        = BarChart()
    grafico.title  = numericas[0]
    grafico.width  = 20
    grafico.height = 14

    dados = Reference(ws_dados, min_col=idx_col, min_row=1, max_row=max_linha)
    grafico.add_data(dados, titles_from_data=True)
    ws_grafico.add_chart(grafico, "B2")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Converte CSV para Excel formatado.")
    parser.add_argument("csv",     help="Arquivo CSV de entrada")
    parser.add_argument("--saida", default=None, help="Nome do arquivo de saída (sem extensão)")
    args = parser.parse_args()

    if not Path(args.csv).exists():
        print(f"  ❌ Arquivo não encontrado: {args.csv}")
        sys.exit(1)

    converter(args.csv, args.saida)
